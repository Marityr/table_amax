import os
import requests
import json
import openpyxl

from openpyxl import Workbook, load_workbook
from dotenv import load_dotenv

load_dotenv()


class Exelfile:
    """Экспорт данных ексель из файла"""

    def read_file_brand() -> list:
        """считывание данных из файла"""

        sheet_list = load_workbook('export.xlsx')
        brand = sheet_list["brand"]

        brand_list = list()
        for cell in brand['A']:
            brand_list.append(cell.value)
 
        return brand_list

    def read_file_sheet():
        sheet_list = load_workbook('export.xlsx')
        sheet = sheet_list["sheet"]

        #tmp = sheet.cell(row=rowit, column=3).value
        number_list = list()
        item = 0
        for cell in sheet['E']:
            number_list.append(cell.value)
            item += 1
            # if item >= 100:
            #     break

        return number_list

    def records():
        """Обработка данных"""
    
        shet_list = load_workbook('export.xlsx')
        sheet = shet_list['sheet']

        iter = 0
        for cell in sheet['E']:
            iter += 1
            data = API_set.get_api(cell.value)

            for item in data:
                print(data[item]['number'])

            if iter >= 2:
                break


class API_set:
    """Взаимодействие с API сервиса"""

    def connect_api(number_list):
        """Подключение по API и получение json"""

        HOST_API = os.getenv('HOST_API')
        USER_API = os.getenv('USER_API')
        PASSWORD_API = os.getenv('PASSWORD_API')

        listen = list()
        items = 0
        for i in number_list:
            conect_url = f"https://{HOST_API}/search/brands/?userlogin={USER_API}&userpsw={PASSWORD_API}&number={i}"
            response = requests.get(conect_url)
            datajson = response.json()
            listen.append(datajson)
            items += 1
            print(items)
        return listen

    def get_api(number):
        """Запрос данных по API"""

        HOST_API = os.getenv('HOST_API')
        USER_API = os.getenv('USER_API')
        PASSWORD_API = os.getenv('PASSWORD_API')   

        conect_url = f"https://{HOST_API}/search/brands/?userlogin={USER_API}&userpsw={PASSWORD_API}&number={number}"
        response = requests.get(conect_url)
        datajson = response.json()

        return datajson




def max_row() -> int:
    """максимальное количество строк с данными"""

    wb = openpyxl.load_workbook('export.xlsx')
    sheet = wb['sheet']
    nb_row = sheet.max_row
    return int(nb_row)

#listen = Exelfile.read_file_sheet()

# listen.pop(0)
#print(listen)
#tempel = API_set.connect_api(listen)
# with open('number_list_two.txt', 'w') as outfile:
#    json.dump(tempel, outfile)
#print(tempel)
# print(len(API_set.connect_api()))
# for i in range(2, 10):
#     print(len(API_set.connect_api(i)))

brand_list = Exelfile.read_file_brand()

Exelfile.records()

# with open('number_list.txt') as json_file:
#     data = json.load(json_file)

# for item in data:
#     for i in item:
#         print(item[i]['brand'])



