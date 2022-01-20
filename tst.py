import os
import requests
import json
import openpyxl

from openpyxl import Workbook, load_workbook
from dotenv import load_dotenv

load_dotenv()


def get_api(number):
    """Запрос данных по API"""

    HOST_API = os.getenv('HOST_API')
    USER_API = os.getenv('USER_API')
    PASSWORD_API = os.getenv('PASSWORD_API')

    conect_url = f"https://{HOST_API}/search/brands/?userlogin={USER_API}&userpsw={PASSWORD_API}&number={number}&useOnlineStocks=1"
    response = requests.get(conect_url)
    datajson = response.json()

    return datajson


def read_file_brand() -> list:
    """считывание данных из файла"""

    sheet_list = load_workbook('export.xlsx')
    brand = sheet_list["brand"]

    brand_list = list()
    for cell in brand['A']:
        brand_list.append(cell.value)

    return brand_list


def records():
    """Обработка данных"""

    shet_list = load_workbook('export_new.xlsx')
    sheet = shet_list['sheet']
    sheet = shet_list.active

    brand_list = read_file_brand()

    iter = 0
    col = 5
    for i in range(2, sheet.max_row):
        iter += 1
        number = sheet.cell(row=i, column=col).value
        data_list = get_api(number)

        col_ls = 6
        for element in data_list:
            print(iter, col_ls, data_list[element])
            if element == 'errorCode' or element == 'errorMessage':
                break
            if data_list[element]['brand'] in brand_list:
                sheet.cell(row=i, column=col_ls).value = data_list[element]['brand']
                sheet.cell(row=i, column=col_ls+1).value = data_list[element]['description']
                col_ls += 2
        
        # ограничение для тестов
        if iter == 10000000:
            break
    shet_list.save('export_new.xlsx')
    shet_list.close()

records()