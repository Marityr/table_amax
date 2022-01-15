import os
import requests
import json
import openpyxl

from openpyxl import Workbook, load_workbook

def read_file_brand() -> list:
        """считывание данных из файла"""

        sheet_list = load_workbook('export.xlsx')
        brand = sheet_list["brand"]

        brand_list = list()
        for cell in brand['A']:
            brand_list.append(cell.value)

        return brand_list

def list_data(number):
    """поиск по номеру из массива"""

    file_path = r'data_product.json'
    with open(file_path, 'r', encoding='utf-8') as file:
        data = json.load(file)

    listent = list()
    for element in data:
        for key, value in element.items():
            a = [value['number'], value['brand'], value['description']] if isinstance(value, dict) else ['none']
            tmp = list()
            if number == a[0]:
                tmp.append(a[0])
                tmp.append(a[1])
                tmp.append(a[2])
                listent.append(tmp)

    return listent  

def records():
    """Обработка данных"""
    
    shet_list = load_workbook('export.xlsx')
    sheet = shet_list['sheet']

    #brand_list = read_file_brand() 

    iter = 0
    for i in range(2, sheet.max_row):
        iter += 1
        number = sheet.cell(row=i, column=5).value
        data_list = list_data(number)
        
        print(iter, data_list)
     
records()